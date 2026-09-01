#!/usr/bin/env python3
"""
verify_form_sync.py
===================
Fails loudly when the Kobo XLSForm and fetch_rdt_results.py drift apart.

WHY THIS EXISTS
---------------
The same content is written down twice: once in the XLSForm (what participants
actually see and what Kobo scores with) and once in fetch_rdt_results.py (what
the dashboard labels and re-scores with). Nothing enforces that they agree, so a
question reworded in one place, or an answer key edited on one side only, would
silently produce a dashboard that mislabels or misgrades real results.

This script compares the two and exits non-zero on any mismatch:

  1. Answer key      - XLSForm `if(${qN}='X',1,0)` calculations vs ANSWER_KEY
  2. Max score       - the total_raw sum and `div N` vs MAX_SCORE
  3. Pass mark       - the result_label threshold vs PASS_THRESHOLD
  4. Question text   - each qN label vs QUESTION_LABELS
  5. Option text     - each qN choice list, in a/b/c/d order, vs OPTION_TEXTS
  6. Province labels - the `province` choice list vs PROVINCE_LABELS
  7. Choice names    - position and test_type names vs POSITION_LABELS / TESTTYPE_LABELS

USAGE
-----
    pip install openpyxl
    python verify_form_sync.py CNM_RDT_PrePostTest_KoboXLSForm_KhmerFixed.xlsx

Run it after editing either side, and before deploying the form to Kobo.
"""

import argparse
import importlib.util
import os
import re
import sys

try:
    import openpyxl
except ImportError:
    sys.exit("openpyxl is required:  pip install openpyxl")


def load_script(path):
    spec = importlib.util.spec_from_file_location("rdt", path)
    mod = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(mod)
    return mod


def sheet_rows(ws):
    """rows as dicts keyed by header name"""
    hdr = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    out = []
    for r in range(2, ws.max_row + 1):
        out.append({h: ws.cell(r, c + 1).value for c, h in enumerate(hdr) if h})
    return out


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("xlsx", help="the Kobo XLSForm to check against")
    ap.add_argument("--script", default="fetch_rdt_results.py")
    args = ap.parse_args()

    for p in (args.xlsx, args.script):
        if not os.path.exists(p):
            sys.exit("not found: %s" % p)

    m = load_script(args.script)
    wb = openpyxl.load_workbook(args.xlsx)
    survey = sheet_rows(wb["survey"])
    choices = sheet_rows(wb["choices"])

    problems = []
    checks = 0

    def check(ok, msg):
        nonlocal checks
        checks += 1
        if not ok:
            problems.append(msg)

    # ---- 1. answer key ----------------------------------------------------
    form_key = {}
    for row in survey:
        calc = row.get("calculation") or ""
        hit = re.fullmatch(r"if\(\$\{(q\d+)\}='([a-d])',1,0\)", str(calc).strip())
        if hit:
            form_key[hit.group(1)] = hit.group(2)

    check(len(form_key) == len(m.ANSWER_KEY),
          "answer key size: XLSForm has %d scored questions, script has %d"
          % (len(form_key), len(m.ANSWER_KEY)))
    for q in sorted(set(form_key) | set(m.ANSWER_KEY), key=lambda s: int(s[1:])):
        want = m.ANSWER_KEY.get(q, (None,))[0]
        got = form_key.get(q)
        check(want == got, "answer key %s: XLSForm=%r script=%r" % (q, got, want))

    # ---- 2. max score -----------------------------------------------------
    total_raw = next((r for r in survey if r.get("name") == "total_raw"), None)
    check(total_raw is not None, "total_raw calculate field missing from XLSForm")
    if total_raw:
        n_terms = len(re.findall(r"\$\{q\d+_pts\}", str(total_raw.get("calculation"))))
        check(n_terms == m.MAX_SCORE,
              "total_raw sums %d questions but MAX_SCORE is %d" % (n_terms, m.MAX_SCORE))
    total_100 = next((r for r in survey if r.get("name") == "total_100"), None)
    if total_100:
        div = re.search(r"div\s+(\d+)", str(total_100.get("calculation")))
        check(div and int(div.group(1)) == m.MAX_SCORE,
              "total_100 divides by %s but MAX_SCORE is %d"
              % (div.group(1) if div else "?", m.MAX_SCORE))

    # ---- 3. pass mark -----------------------------------------------------
    label_row = next((r for r in survey if r.get("name") == "result_label"), None)
    if label_row:
        thr = re.search(r">=\s*([\d.]+)", str(label_row.get("calculation")))
        check(thr and float(thr.group(1)) == float(m.PASS_THRESHOLD),
              "result_label threshold is %s but PASS_THRESHOLD is %s"
              % (thr.group(1) if thr else "?", m.PASS_THRESHOLD))
    else:
        check(False, "result_label calculate field missing from XLSForm")

    # ---- 4. question text -------------------------------------------------
    for row in survey:
        nm = row.get("name")
        if nm in m.QUESTION_LABELS:
            check(str(row.get("label")) == m.QUESTION_LABELS[nm],
                  "question text %s differs:\n      XLSForm: %s\n      script : %s"
                  % (nm, row.get("label"), m.QUESTION_LABELS[nm]))

    # ---- 5. option text ---------------------------------------------------
    for q, texts in m.OPTION_TEXTS.items():
        opts = [c for c in choices if c.get("list_name") == q]
        by_name = {str(c.get("name")): str(c.get("label")) for c in opts}
        check(len(opts) == len(m.OPTION_LETTERS),
              "%s has %d choices in the XLSForm, script expects %d"
              % (q, len(opts), len(m.OPTION_LETTERS)))
        for i, letter in enumerate(m.OPTION_LETTERS):
            if letter in by_name:
                check(by_name[letter] == texts[i],
                      "option %s.%s differs:\n      XLSForm: %s\n      script : %s"
                      % (q, letter, by_name[letter], texts[i]))
            else:
                check(False, "option %s.%s missing from the XLSForm" % (q, letter))

    # ---- 6. province labels ----------------------------------------------
    provs = {str(c.get("name")): str(c.get("label"))
             for c in choices if c.get("list_name") == "province"}
    check(set(provs) == set(m.PROVINCE_LABELS),
          "province choice names differ: XLSForm=%s script=%s"
          % (sorted(provs), sorted(m.PROVINCE_LABELS)))
    for code, label in sorted(m.PROVINCE_LABELS.items()):
        if code in provs:
            check(provs[code] == label,
                  "province %s differs:\n      XLSForm: %s\n      script : %s"
                  % (code, provs[code], label))
    check(list(m.PROVINCE_ORDER) == sorted(m.PROVINCE_ORDER, key=lambda c: int(c.split("_")[1])),
          "PROVINCE_ORDER is not in roster order: %s" % (m.PROVINCE_ORDER,))

    # ---- 7. choice names the scorer depends on ---------------------------
    for lst, expected, what in (("position", set(m.POSITION_LABELS), "POSITION_LABELS"),
                                ("test_type", set(m.TESTTYPE_LABELS), "TESTTYPE_LABELS")):
        names = {str(c.get("name")) for c in choices if c.get("list_name") == lst}
        check(names == expected,
              "%s choice names %s do not match %s %s"
              % (lst, sorted(names), what, sorted(expected)))

    # ---- report -----------------------------------------------------------
    print("verify_form_sync: %d checks" % checks)
    print("  XLSForm : %s" % args.xlsx)
    print("  script  : %s" % args.script)
    if problems:
        print("\n%d MISMATCH(ES):" % len(problems))
        for p in problems:
            print("  - %s" % p)
        sys.exit(1)
    print("\nOK - the XLSForm and the script agree.")


if __name__ == "__main__":
    main()
